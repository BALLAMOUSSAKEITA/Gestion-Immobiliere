from datetime import date
from math import ceil
from pathlib import Path
from uuid import UUID, uuid4

from fastapi import HTTPException, status
from openpyxl import Workbook
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.pdfgen import canvas
from sqlalchemy.orm import Session, joinedload

from app.core.config import get_settings
from app.models.enums import ReportType
from app.models.report import ReportSnapshot
from app.models.role import Role
from app.models.user import User
from app.schemas.report import ReportDetail, ReportGenerateRequest, ReportListResponse, ReportSummary
from app.services.dashboard_service import DashboardService


class ReportService:
    def __init__(self, db: Session) -> None:
        self.db = db
        settings = get_settings()
        self.upload_dir = Path(settings.upload_dir)
        self.agency_name = settings.agency_name

    def _ensure_access(self, actor: User) -> None:
        if actor.role.code not in ("super_admin", "admin_familial", "proprietaire"):
            raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Accès non autorisé")

    def list_reports(
        self,
        actor: User,
        *,
        page: int = 1,
        page_size: int = 20,
    ) -> ReportListResponse:
        self._ensure_access(actor)
        query = self.db.query(ReportSnapshot).options(joinedload(ReportSnapshot.generator))
        if actor.role.code == "proprietaire":
            query = query.filter(ReportSnapshot.generated_by == actor.id)
        total = query.count()
        records = (
            query.order_by(ReportSnapshot.generated_at.desc())
            .offset((page - 1) * page_size)
            .limit(page_size)
            .all()
        )
        pages = ceil(total / page_size) if total else 0
        return ReportListResponse(
            items=[self._to_summary(record) for record in records],
            total=total,
            page=page,
            page_size=page_size,
            pages=pages,
        )

    def get_report(self, actor: User, report_id: UUID) -> ReportDetail:
        self._ensure_access(actor)
        record = self._get_or_404(report_id)
        if actor.role.code == "proprietaire" and record.generated_by != actor.id:
            raise HTTPException(status_code=403, detail="Accès non autorisé")
        summary = self._to_summary(record)
        return ReportDetail(**summary.model_dump(), data=record.data)

    def generate_report(self, actor: User, payload: ReportGenerateRequest) -> ReportDetail:
        self._ensure_access(actor)
        filters = payload.filters.model_dump() if payload.filters else {}
        building_id = UUID(filters["building_id"]) if filters.get("building_id") else None
        owner_profile_id = (
            UUID(filters["owner_profile_id"]) if filters.get("owner_profile_id") else None
        )

        data = DashboardService(self.db).collect_report_data(
            actor,
            period_start=payload.period_start,
            period_end=payload.period_end,
            building_id=building_id,
            owner_profile_id=owner_profile_id,
        )

        snapshot = ReportSnapshot(
            report_type=payload.report_type,
            period_start=payload.period_start,
            period_end=payload.period_end,
            filters=filters or None,
            data=data,
            generated_by=actor.id,
        )
        self.db.add(snapshot)
        self.db.flush()

        report_id = snapshot.id
        if "pdf" in payload.export_formats:
            snapshot.pdf_url = self._write_pdf(snapshot)
        if "excel" in payload.export_formats:
            snapshot.excel_url = self._write_excel(snapshot)

        self.db.commit()
        return self.get_report(actor, report_id)

    def get_pdf_path(self, actor: User, report_id: UUID) -> Path:
        record = self.get_report(actor, report_id)
        if not record.pdf_url:
            raise HTTPException(status_code=404, detail="PDF non disponible")
        return self._resolve_path(record.pdf_url)

    def get_excel_path(self, actor: User, report_id: UUID) -> Path:
        record = self.get_report(actor, report_id)
        if not record.excel_url:
            raise HTTPException(status_code=404, detail="Excel non disponible")
        return self._resolve_path(record.excel_url)

    def generate_scheduled(self, report_type: ReportType, period_start: date, period_end: date) -> None:
        admin = (
            self.db.query(User)
            .join(Role, User.role_id == Role.id)
            .filter(Role.code == "super_admin", User.is_active.is_(True))
            .first()
        )
        if admin is None:
            return
        payload = ReportGenerateRequest(
            report_type=report_type,
            period_start=period_start,
            period_end=period_end,
            export_formats=["pdf", "excel"],
        )
        self.generate_report(admin, payload)

    def _write_pdf(self, snapshot: ReportSnapshot) -> str:
        target_dir = self.upload_dir / "reports"
        target_dir.mkdir(parents=True, exist_ok=True)
        filename = f"report_{snapshot.id}.pdf"
        path = target_dir / filename

        kpis = snapshot.data.get("kpis", {})
        c = canvas.Canvas(str(path), pagesize=A4)
        width, height = A4
        y = height - 2 * cm

        c.setFont("Helvetica-Bold", 16)
        c.drawString(2 * cm, y, f"{self.agency_name} — Rapport {snapshot.report_type.value}")
        y -= 1 * cm
        c.setFont("Helvetica", 11)
        c.drawString(
            2 * cm,
            y,
            f"Période : {snapshot.period_start} → {snapshot.period_end}",
        )
        y -= 1.5 * cm

        c.setFont("Helvetica-Bold", 13)
        c.drawString(2 * cm, y, "Résumé KPI")
        y -= 0.8 * cm
        c.setFont("Helvetica", 10)
        for key, label in [
            ("total_buildings", "Immeubles"),
            ("occupied_units", "Logements occupés"),
            ("free_units", "Logements libres"),
            ("expected_rent_month", "Loyers attendus"),
            ("collected_rent_month", "Loyers encaissés"),
            ("overdue_amount", "Impayés"),
            ("expenses_month", "Dépenses"),
            ("net_profit_month", "Bénéfice net"),
        ]:
            if key in kpis and kpis[key] is not None:
                c.drawString(2 * cm, y, f"{label} : {kpis[key]}")
                y -= 0.5 * cm

        c.save()
        return f"/uploads/reports/{filename}"

    def _write_excel(self, snapshot: ReportSnapshot) -> str:
        target_dir = self.upload_dir / "reports"
        target_dir.mkdir(parents=True, exist_ok=True)
        filename = f"report_{snapshot.id}.xlsx"
        path = target_dir / filename

        wb = Workbook()
        summary = wb.active
        summary.title = "Résumé"
        summary.append(["Rapport", snapshot.report_type.value])
        summary.append(["Début", str(snapshot.period_start)])
        summary.append(["Fin", str(snapshot.period_end)])

        kpis = snapshot.data.get("kpis", {})
        summary.append([])
        summary.append(["Indicateur", "Valeur"])
        for key, value in kpis.items():
            summary.append([key, value])

        overdues = snapshot.data.get("top_overdues", {}).get("items", [])
        ws_overdue = wb.create_sheet("Impayés")
        ws_overdue.append(["Locataire", "Logement", "Montant", "Jours retard"])
        for item in overdues:
            ws_overdue.append(
                [
                    item.get("tenant_name"),
                    item.get("unit_code"),
                    float(item.get("amount_remaining", 0)),
                    item.get("days_overdue"),
                ]
            )

        expenses = snapshot.data.get("expenses_by_category", {}).get("slices", [])
        ws_exp = wb.create_sheet("Dépenses")
        ws_exp.append(["Catégorie", "Montant", "Nombre"])
        for item in expenses:
            ws_exp.append([item.get("category"), float(item.get("amount", 0)), item.get("count")])

        wb.save(path)
        return f"/uploads/reports/{filename}"

    def _resolve_path(self, file_url: str) -> Path:
        relative = file_url.removeprefix("/uploads/")
        path = self.upload_dir / relative
        if not path.exists():
            raise HTTPException(status_code=404, detail="Fichier introuvable")
        return path

    def _get_or_404(self, report_id: UUID) -> ReportSnapshot:
        record = (
            self.db.query(ReportSnapshot)
            .options(joinedload(ReportSnapshot.generator))
            .filter(ReportSnapshot.id == report_id)
            .first()
        )
        if record is None:
            raise HTTPException(status_code=404, detail="Rapport introuvable")
        return record

    def _to_summary(self, record: ReportSnapshot) -> ReportSummary:
        generator_name = None
        if record.generator:
            generator_name = f"{record.generator.first_name} {record.generator.last_name}"
        return ReportSummary(
            id=str(record.id),
            report_type=record.report_type,
            period_start=record.period_start,
            period_end=record.period_end,
            filters=record.filters,
            pdf_url=record.pdf_url,
            excel_url=record.excel_url,
            generated_by_name=generator_name,
            generated_at=record.generated_at,
        )
